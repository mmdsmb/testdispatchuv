import hashlib
from math import radians, sin, cos, sqrt, atan2
import pandas as pd
import numpy as np
import os
from datetime import datetime
import logging
from typing import Optional, Dict
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from datetime import datetime
from dotenv import load_dotenv
import json
import base64
from googleapiclient.discovery import build
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from app.db.postgres import PostgresDataSource

# # Configuration des scopes pour Google Drive API
SCOPES = ['https://www.googleapis.com/auth/drive.file']

logger = logging.getLogger(__name__)

def generate_address_hash(address: str) -> str:
    """Génère un hash MD5 standardisé pour les adresses"""
    return hashlib.md5(address.strip().lower().encode('utf-8')).hexdigest()

def generate_route_hash(pickup_address: str, destination_address: str) -> str:
    """Génère un hash unique pour un trajet (combinaison des deux adresses)"""
    combined = f"{pickup_address.strip().lower()}_{destination_address.strip().lower()}"
    return generate_address_hash(combined)

def calculate_bird_distance(origin_lat: float, origin_lng: float, 
                          dest_lat: float, dest_lng: float) -> float:
    """Calcule la distance à vol d'oiseau en km (formule Haversine)"""
    R = 6371  # Rayon de la Terre en km
    lat1, lon1, lat2, lon2 = map(radians, [origin_lat, origin_lng, dest_lat, dest_lng])
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    return round(R * c, 2)

def convertir_si_nombre(nombre):
    """
    nombre = 'Paris' #1234.456  # ou 'Paris'
    nombre_sans_decimal = convertir_si_nombre(nombre)
    print(nombre_sans_decimal)
    """
    try:
        return int(float(nombre))
    except ValueError:
        return nombre

def ajoutColonne(df, nom_colonne, valeur_colonne, idx=None):
    """
    Ajoute une nouvelle colonne à un DataFrame Pandas.

    Args:
        df (DataFrame): Le DataFrame auquel ajouter la colonne.
        nom_colonne (str): Le nom de la nouvelle colonne.
        valeur_colonne (list): Les valeurs de la nouvelle colonne.
        idx (int, optional): L'indice d'insertion de la colonne. 
        Par défaut, la colonne est ajoutée à la fin.

    Returns:
        DataFrame: Le DataFrame avec la nouvelle colonne ajoutée.
    """
    if idx is None:
        # Par défaut, l'indice est défini sur le nombre de colonnes actuelles
        idx = df.shape[1]  
    df.insert(loc=idx, column=nom_colonne, value=valeur_colonne)
    return df

def Valeur_manquante(df, colonne_obligatoire):
    """
    Ajoute une colonne au DataFrame indiquant les colonnes contenant des valeurs manquantes pour chaque ligne.

    Args:
        df (DataFrame): Le DataFrame à traiter.
        colonne_obligatoire (list): Liste des colonnes à vérifier pour les valeurs manquantes.

    Returns:
        DataFrame: Le DataFrame avec une colonne supplémentaire 'ERREURS' indiquant les colonnes manquantes pour chaque ligne.
    """
    # Sélectionner les lignes avec des valeurs manquantes dans les colonnes spécifiées
    missing_columns = df[df[colonne_obligatoire].isnull().any(axis=1)]
    
    # Ajouter une colonne 'ERREURS' pour indiquer les colonnes manquantes
    missing_columns['ERREURS'] = missing_columns[colonne_obligatoire].apply(lambda row: ', '.join(row.index[row.isnull()]), axis=1)
    
    # Ajouter la colonne 'ERREURS' au DataFrame original
    df = ajoutColonne(df, 'ERREURS', missing_columns['ERREURS'], df.shape[1])
    
    return df



def verifier_date(date):
    """
    Vérifie si la date est au format jj-mm-aaaa.

    Args:
        date (str): La date à vérifier.

    Returns:
        str: "OK" si la date est au format correct, sinon "KO".
    """
    try:
        pd.to_datetime(date, format='%d-%m-%Y')
        return "OK"
    except ValueError:
        return "KO"

def verifier_format_date(df, colonne):
    """
    Vérifie le format des colonnes de dates dans un DataFrame.

    Args:
        df (DataFrame): Le DataFrame contenant les données.
        colonne (list): La liste des noms des colonnes contenant les dates à vérifier.

    Returns:
        DataFrame: Le DataFrame avec des colonnes supplémentaires indiquant si les dates sont au bon format.
    """
    for col in colonne:
        verification_colonne = df[col].apply(verifier_date)
        df.insert(df.columns.get_loc(col) + 1, colonne + '_Verification', verification_colonne)
    return df

def convert_date_heure(row,champ_date:str, champ_heure:str):
    date_str = row[champ_date]
    heure_str = row[champ_heure]
    # Convertir les dates et heures en datetime
    # date_obj = pd.to_datetime(date_str, format='%d/%m/%Y', errors='coerce')
    # heure_obj = pd.to_datetime(heure_str, format='%Hh%M', errors='coerce').dt.time
    # date_obj = pd.to_datetime(date_str).dt.strftime('%d/%m/%Y')
    # heure_obj = pd.to_datetime(heure_str).dt.strftime('%H:%M')

    #if pd.isnull(date_obj) or pd.isnull(heure_obj):
    if pd.isnull(date_str) or pd.isnull(heure_str):
        return None
    else:
        # Combiner les dates et les heures
        # Concaténer la colonne de date et la colonne d'heure
        #date_heure_combinee = pd.to_datetime((date_str) + ' ' + str(heure_str))
        date_heure_combinee = date_str + ' ' + heure_str        

    return date_heure_combinee
    # return pd.Timestamp.combine(date_obj, heure_obj)




def format_heure(heure_str):
    # Debug initial
    print(f"Input: {heure_str} (type: {type(heure_str)})")
    
    # Gestion des valeurs manquantes
    if pd.isna(heure_str) or str(heure_str).strip() in ['', 'nan', 'None']:
        return ""
    
    try:
        heure_str = str(heure_str)  # Conversion forcée en string
        print(f"conversion forcée: {heure_str} (type: {type(heure_str)})")

        # Nettoyage supplémentaire
        heure_str = heure_str.strip().replace(' ', '')
        print(f"nettoyage supplémentaire: {heure_str} (type: {type(heure_str)})")
        # Gestion du format "07h50"
        if 'h' in heure_str:
            parts = heure_str.split('h')
            if len(parts) == 2:
                h = parts[0].zfill(2)  # Ajoute un 0 devant si nécessaire
                m = parts[1][:2].ljust(2, '0')  # Prend les 2 premiers caractères et complète avec 0
                print(f"format 07h50: {h}:{m}")
                return f"{h}:{m}"
        
        # Autres formats possibles à gérer ici...
        
    except Exception as e:
        print(f"Erreur de conversion pour '{heure_str}': {str(e)}")
        return ""
    
    return ""  # Retour par défaut

def verifier_champs_manquants(row, champs_a_verifier):
    """Vérifie les champs manquants pour une ligne donnée."""
    champs_manquants = []
    for champ in champs_a_verifier:
        # Vérifier si le champ est manquant ou vide
        if pd.isna(row[champ]) or str(row[champ]).strip() == '':
            champs_manquants.append(champ)
        # Vérifier spécifiquement les erreurs de conversion pour les champs numériques
        elif champ in ['ID', 'Nombre-prs-AR', 'Nombre-prs-Ret'] and f"{champ}_erreur" in row and row[f"{champ}_erreur"]:
            champs_manquants.append(f"{champ}_invalide")
    
    return ', '.join(champs_manquants) if champs_manquants else ''

def clean_integer(value):
    """Nettoie et valide une valeur devant être un entier."""
    if pd.isna(value) or value == '':
        return None, "Valeur manquante"
    
    try:
        # Supprimer les espaces et caractères non numériques
        cleaned = ''.join(c for c in str(value) if c.isdigit())
        if not cleaned:
            return None, "Valeur non numérique"
        
        return int(cleaned), None
    except Exception as e:
        return None, f"Erreur conversion: {str(e)}"

def decode_base64_credentials(encoded_credentials):
    """Décode les credentials encodés en base64."""
    try:
        # Décoder la chaîne base64
        decoded_bytes = base64.b64decode(encoded_credentials)
        # Convertir en chaîne de caractères
        decoded_str = decoded_bytes.decode('utf-8')
        # Parser le JSON
        return json.loads(decoded_str)
    except Exception as e:
        raise ValueError(f"Erreur lors du décodage des credentials: {str(e)}")
    


def get_google_drive_service():
    """Version pour Service Account"""
    encoded_credentials = os.getenv('GOOGLE_CREDENTIALS_BASE64')
    if not encoded_credentials:
        raise ValueError("GOOGLE_CREDENTIALS_BASE64 manquant")
    
    credentials_dict = decode_base64_credentials(encoded_credentials)
    
    if 'type' not in credentials_dict or credentials_dict['type'] != 'service_account':
        raise ValueError("Les identifiants doivent être pour un Service Account")
    
    creds = service_account.Credentials.from_service_account_info(
        credentials_dict,
        scopes=SCOPES
    )
    
    return build('drive', 'v3', credentials=creds)




def upload_to_drive(file_path, folder_id):
    """Upload un fichier vers Google Drive dans le dossier spécifié."""
    service = get_google_drive_service()
    
    file_metadata = {
        'name': os.path.basename(file_path),
        'parents': [folder_id]
    }
    
    media = MediaFileUpload(file_path, resumable=True)
    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()
    
    return file.get('id')


# async def save_and_upload_to_drive(
#     df: pd.DataFrame,
#     folder_id: str,
#     file_prefix: str = "rapport",
#     file_extension: str = "xlsx",
#     drive_service=None,
#     local_cleanup: bool = True
# ) -> Optional[str]:
#     """
#     Enregistre un DataFrame dans un fichier Excel et l'upload sur Google Drive
    
#     Args:
#         df: DataFrame à enregistrer
#         folder_id: ID du dossier Google Drive destination
#         file_prefix: Préfixe du nom de fichier
#         file_extension: Extension du fichier (xlsx ou csv)
#         drive_service: Service Google Drive initialisé
#         local_cleanup: Supprimer le fichier local après upload
    
#     Returns:
#         str: ID du fichier sur Google Drive ou None en cas d'erreur
#     """
#     try:
#         # Création du nom de fichier avec timestamp
#         timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#         filename = f"{file_prefix}_{timestamp}.{file_extension}"
        
#         # Enregistrement local
#         if file_extension == 'xlsx':
#             df.to_excel(filename, index=False)
#         elif file_extension == 'csv':
#             df.to_csv(filename, index=False)
#         else:
#             raise ValueError("Extension de fichier non supportée. Utiliser 'xlsx' ou 'csv'")
        
#         logger.info(f"Fichier {filename} créé localement")
        
#         if not drive_service:
#             drive_service = get_google_drive_service()
        
#         # Upload vers Google Drive
#         file_metadata = {
#             'name': filename,
#             'parents': [folder_id]
#         }
        
#         media = MediaFileUpload(filename, 
#                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' 
#                               if file_extension == 'xlsx' else 'text/csv')
        
#         file = drive_service.files().create(
#             body=file_metadata,
#             media_body=media,
#             fields='id'
#         ).execute()
        
#         file_id = file.get('id')
#         logger.info(f"Fichier uploadé avec succès sur Drive. ID: {file_id}")
        
#         # Nettoyage du fichier local si demandé
#         if local_cleanup:
#             os.remove(filename)
#             logger.info(f"Fichier local {filename} supprimé")
        
#         return file_id
        
#     except Exception as e:
#         logger.error(f"Erreur lors de l'enregistrement/upload du fichier: {str(e)}", exc_info=True)
#         return None

def load_to_excel(df_to_save, file_path, format_excel=False ,index=False):

    """
    prend en paramètre le DataFrame df_to_save à enregistrer et le chemin du fichier file_path où enregistrer le fichier. 
    Elle prend également un paramètre format_excel pour indiquer 
    si vous voulez formater le fichier Excel avec une alternance de couleurs de ligne et un entête en gras.
    """
    # Créer un writer Excel avec pandas
    writer = pd.ExcelWriter(file_path, engine='openpyxl')

    # Écrire le DataFrame dans le fichier Excel
    #df_to_save.to_excel(writer, sheet_name='Sheet1', index=False)
    df_to_save.to_excel(writer, sheet_name='Sheet1', index=index)

    if format_excel:
        from openpyxl.styles import Font, PatternFill

        # Accéder à la feuille de calcul
        worksheet = writer.sheets['Sheet1']

        # Appliquer une police grasse à l'entête
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))


        # Appliquer une alternance de couleurs de ligne et ajouter des bordures
        for r_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                if r_idx % 2 == 0:
                    fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                else:
                    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.fill = fill
                cell.border = thin_border  # Ajouter des bordures

        # Ajuster la largeur des colonnes et ajouter des bordures
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                cell.border = thin_border  # Ajouter des bordures
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


    # Enregistrer le fichier Excel
    #writer.save()
    writer.close()

async def save_and_upload_to_drive(
    df: pd.DataFrame,
    folder_id: str,
    file_prefix: str = "rapport",
    file_extension: str = "xlsx",
    drive_service=None,
    local_cleanup: bool = True,
    subfolder_name: Optional[str] = None,
    format_excel: bool = False,
    index: bool = False
) -> Optional[str]:
    """
    Enregistre un DataFrame dans un fichier Excel (avec formatage optionnel) et l'upload dans un sous-dossier Google Drive
    
    Args:
        df: DataFrame à enregistrer
        folder_id: ID du dossier Google Drive parent
        file_prefix: Préfixe du nom de fichier
        file_extension: Extension du fichier (xlsx uniquement pour le formatage)
        drive_service: Service Google Drive initialisé
        local_cleanup: Supprimer le fichier local après upload
        subfolder_name: Nom du sous-dossier (créé s'il n'existe pas)
        format_excel: Appliquer un formatage avancé (alternance de couleurs, bordures, etc.)
        index: Inclure l'index du DataFrame dans l'export
    
    Returns:
        str: ID du fichier sur Google Drive ou None en cas d'erreur
    """
    try:
        # Validation du DataFrame
        if df.empty:
            logger.warning("Le DataFrame est vide - aucun fichier créé")
            return None

        # Création du nom de fichier avec timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{file_prefix}.{file_extension}"
        
        # Enregistrement local avec la fonction personnalisée
        if file_extension == 'xlsx':
            load_to_excel(
                df_to_save=df,
                file_path=filename,
                format_excel=format_excel,
                index=index
            )
        elif file_extension == 'csv':
            df.to_csv(filename, index=index)
        else:
            raise ValueError("Extension non supportée. Utiliser 'xlsx' ou 'csv'")
        
        logger.info(f"Fichier {filename} créé localement avec {'formatage' if format_excel else 'sans formatage'}")

        # Initialisation du service Drive si non fourni
        if not drive_service:
            drive_service = get_google_drive_service()

        # Gestion du sous-dossier
        target_folder_id = folder_id
        if subfolder_name:
            # Vérifier si le sous-dossier existe déjà
            query = f"name='{subfolder_name}' and mimeType='application/vnd.google-apps.folder' and '{folder_id}' in parents and trashed=false"
            results = drive_service.files().list(
                q=query,
                fields="files(id, name)"
            ).execute()
            folders = results.get('files', [])
            
            if folders:
                # Sous-dossier existe déjà
                target_folder_id = folders[0]['id']
                logger.info(f"Sous-dossier existant utilisé: {subfolder_name} (ID: {target_folder_id})")
            else:
                # Créer le nouveau sous-dossier
                folder_metadata = {
                    'name': subfolder_name,
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [folder_id]
                }
                folder = drive_service.files().create(
                    body=folder_metadata,
                    fields='id'
                ).execute()
                target_folder_id = folder.get('id')
                logger.info(f"Sous-dossier créé: {subfolder_name} (ID: {target_folder_id})")

        # Upload vers Google Drive
        file_metadata = {
            'name': filename,
            'parents': [target_folder_id]
        }
        
        media = MediaFileUpload(
            filename, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' 
            if file_extension == 'xlsx' else 'text/csv'
        )
        
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        file_id = file.get('id')
        logger.info(f"Fichier uploadé avec succès dans {'sous-dossier ' + subfolder_name if subfolder_name else 'dossier principal'}. ID: {file_id}")
        
        # Nettoyage
        if local_cleanup and os.path.exists(filename):
            os.remove(filename)
            logger.info(f"Fichier local {filename} supprimé")
        
        return file_id
        
    except Exception as e:
        logger.error(f"Erreur lors de l'enregistrement/upload: {str(e)}", exc_info=True)
        if 'filename' in locals() and os.path.exists(filename):
            os.remove(filename)
        return None

async def extract_data_from_query(query: str) -> pd.DataFrame:
    """
    Exécute une requête SQL et retourne un DataFrame pandas avec les colonnes de la table.
    Gère automatiquement la connexion et la déconnexion à la base de données.

    Args:
        query (str): La requête SQL à exécuter.

    Returns:
        pd.DataFrame: Un DataFrame contenant les résultats de la requête avec les noms de colonnes.
    """
    db_source = PostgresDataSource()
    try:
        # Établir la connexion
        await db_source.connect()
        
        # Exécuter la requête et récupérer les résultats avec les noms de colonnes
        async with db_source.conn.cursor() as cursor:
            await cursor.execute(query)
            
            # Récupérer les noms des colonnes
            column_names = [desc[0] for desc in cursor.description]
            
            # Récupérer les données
            results = await cursor.fetchall()
        
        # Convertir les résultats en DataFrame pandas avec les noms de colonnes
        df = pd.DataFrame(results, columns=column_names)
        
        return df
    
    except Exception as e:
        logger.error(f"Erreur lors de l'extraction des données: {str(e)}", exc_info=True)
        raise
    finally:
        # Fermer la connexion dans tous les cas (succès ou erreur)
        await db_source.disconnect()
        
# async def save_and_upload_to_drive(df, folder_id, file_prefix, subfolder_name, format_excel, index):
#     """Sauvegarde et upload un DataFrame vers Google Drive en créant le sous-dossier si nécessaire"""
#     try:
#         # 1. Vérifier/créer le sous-dossier
#         subfolder_id = await create_or_get_subfolder(folder_id, subfolder_name)
        
#         # 2. Sauvegarder le fichier Excel temporaire
#         output_file = f"{file_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         df.to_excel(output_file, index=index)
        
#         # 3. Upload vers le sous-dossier (implémentez votre logique d'upload ici)
#         # Exemple fictif :
#         file_id = await upload_to_drive(output_file, subfolder_id)
#         return file_id
        
#     except Exception as e:
#         logger.error(f"Erreur lors de l'upload vers Drive: {str(e)}")
#         raise

# async def create_or_get_subfolder(parent_id, folder_name):
#     """Crée ou récupère un sous-dossier dans Google Drive"""
#     # Implémentez la logique de création/récupération du sous-dossier
#     # Exemple fictif :
#     existing_folder = await check_folder_exists(parent_id, folder_name)
#     if existing_folder:
#         return existing_folder['id']
#     else:
#         return await create_folder(parent_id, folder_name)

# async def create_folder(parent_id: str, folder_name: str) -> str:
#     """Crée un dossier dans Google Drive et retourne son ID"""
#     service = get_google_drive_service()
#     folder_metadata = {
#         'name': folder_name,
#         'mimeType': 'application/vnd.google-apps.folder',
#         'parents': [parent_id]
#     }
#     folder = service.files().create(
#         body=folder_metadata,
#         fields='id'
#     ).execute()
#     return folder.get('id')

# async def check_folder_exists(parent_id: str, folder_name: str) -> Optional[Dict]:
#     """Vérifie si un dossier existe déjà dans Google Drive"""
#     service = get_google_drive_service()
#     query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed=false"
#     results = service.files().list(
#         q=query,
#         fields="files(id, name)"
#     ).execute()
#     folders = results.get('files', [])
#     return folders[0] if folders else None